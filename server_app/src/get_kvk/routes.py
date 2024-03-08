import pandas as pd
from fastapi import APIRouter, Depends, Request, Form
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from starlette import status
from starlette.responses import RedirectResponse, FileResponse

from server_app.database.database import get_async_session
from server_app.src.auth.model import User, department, post, risk, executor_group, control_group, User, zamruk
from server_app.src.utils import current_active_user

import openpyxl as opx

router = APIRouter(
    prefix='/kvk',
    tags=['kvk'],
    dependencies=[Depends(current_active_user)]
)


async def select_risks(dep_id: int, session: AsyncSession = Depends(get_async_session)) -> pd.DataFrame:
    """
    dep_id: Департамент, по которому нужно открыть риски
    return: 3 датафрейма с рисками, контролирующими, исполнителями

    """

    # Риски
    query = select(risk.id, risk.name, department.name)
    query = query.join(department, risk.fk_dep == department.id)
    query = query.where(department.id == dep_id)
    result = await session.execute(query)
    risk_dict = []
    list_currentRisk = []
    for row in result:
        risk_dict.append({'id': row[0], 'name': row[1], 'dep_name': row[2]})
        list_currentRisk.append(row[0])

    # Сотрудники в группе исполнителей
    query = select(risk.id, executor_group.fk_post_id, post.name, User.id, User.username)
    query = query.join(risk, executor_group.fk_risk_id == risk.id)
    query = query.join(post, executor_group.fk_post_id == post.id)
    query = query.join(User, post.id == User.fk_post_id)
    query = query.filter(risk.id.in_(list_currentRisk))
    query = query.where(User.fk_department_id == dep_id)
    query = query.order_by(executor_group.fk_post_id, User.username)
    result = await session.execute(query)
    executors_dict = []
    for row in result:
        executors_dict.append({'id': row[0], 'executor_post_id': row[1], 'executor_post_name': row[2],
                               'user_id': row[3], 'user_name': row[4]})



    query = select(zamruk.fk_user_id, User.username, User.fk_post_id, post.name)
    query = query.join(User, zamruk.fk_user_id == User.id)
    query = query.join(post, User.fk_post_id == post.id)
    query = query.where(zamruk.fk_dep_id == dep_id)
    result = await session.execute(query)
    result = result.first()
    kurator = {
        'username': result[1],
        'post_id': result[2],
        'post_name': result[3]
    }

    query = select(User.username, User.fk_post_id, post.name)
    query = query.join(post, User.fk_post_id == post.id)
    query = query.where(User.fk_department_id == dep_id)
    query = query.order_by(User.fk_post_id, User.username)
    query = query.limit(1)
    result = await session.execute(query)
    result = result.first()
    head_dep = {
        'username': result[0],
        'post_id': result[1],
        'post_name': result[2]
    }


    df_executor = pd.DataFrame(executors_dict)
    df_risk = pd.DataFrame(risk_dict)

    return df_risk, df_executor, head_dep, kurator


def create_kvk_data(df_risk, df_executor, head_department, kurator, ws):
    """
    Заполнить квк по
    """
    # переменные
    department_name = df_risk['dep_name'].iloc[0]
    today = 'Прошлый век сегодняшнего столетия'
    header_table = ['№п/п', 'Предмет внутреннего контроля',
                    'ФИО, должность должностного лица, ответственного за выполнение операции (действия)',
                    'Периодичность выполнения операции (действия)',
                    'ФИО, должности должностных лиц, осуществляющих контрольные действия',
                    'Метод контроля', 'Способ контроля', 'Периодичность контрольных действий',
                    'Подписи должностных лиц, осуществ-ляющих контрольные действия']
    col_mapping = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F', 6: 'G', 7: 'H', 8: 'I'}

    # Отрисовка шапки документа
    ws['A1'] = "Экземпляр №"
    ws['F1'] = f"«УТВЕРЖДАЮ»\n" \
               f"\n{kurator['post_name']}" \
               f"\nФедерального казначейства\n" \
               f"\n_________________ {head_department['username']}\n" \
               f"\n«_______»_________________{today} г."
    ws['A3'] = f"Карта внутреннего контроля на {today} год"
    ws['A5'] = 'Наименование органа Федерального казначейства, казенного учреждения: Межрегиональное бухгалтерское управление Федерального казначейства'
    ws['A6'] = f"Наименование структурного подразделения: {department_name}"

    # Шапка табилцы
    for col in range(len(header_table)):
        row_header = '8'
        row_number = '9'
        ws[col_mapping[col] + row_header] = header_table[col]
        ws[col_mapping[col] + row_number] = col + 1

    #Заполнение таблицы
    #переменные
    current_row = 10
    current_risk = -1

    while current_risk < len(df_risk):
        current_risk += 1
        current_risk_name = df_risk.loc[df_risk.id == current_risk, 'name'].to_string(index=False)
        executors_in_risk = df_executor.loc[df_executor['id'] == current_risk]
        row_len = len(executors_in_risk) + 1
        counter = 1
        if row_len == 2:
            controller = f'{kurator}, заместитель руководителя Управления'
        else:
            controller = f'{head_department["username"]}, начальник отдела'

        df_executors_to_excel = executors_in_risk['user_name'] +', ' + executors_in_risk['executor_post_name'].str.lower()+ ';'
        df_executors_to_excel = df_executors_to_excel.tolist()
        df_executors_to_excel = '\n'.join(df_executors_to_excel)
        df_executors_to_excel = df_executors_to_excel[:len(df_executors_to_excel)-1] + '.'

        #заполнение строк
        ws['A' + str(current_row)] = counter
        ws['B' + str(current_row)] = current_risk_name
        ws['C' + str(current_row)] = df_executors_to_excel
        ws['D' + str(current_row)] = 'В установленные сроки'
        ws['E' + str(current_row)] = df_executors_to_excel
        ws['F' + str(current_row)] = 'Самоконтроль'
        ws['G' + str(current_row)] = 'Сплошной'
        ws['H' + str(current_row)] = 'По мере совершения операции'

        ws['E' + str((current_row+row_len-1))] = controller
        ws['F' + str((current_row+row_len-1))] = 'Контроль по уровню подчиненности '
        ws['G' + str((current_row+row_len-1))] = 'Сплошной'
        ws['H' + str((current_row+row_len-1))] = 'По мере совершения операции'




def create_kvk_style():
    """"""
    pass


def make_kvk(df_risk, df_executor, head_department, kurator) -> int:
    """Создает документ, возвращает ссылку путь до этого документа (название)"""
    wb = opx.Workbook()
    ws = wb.active
    create_kvk_data(df_risk, df_executor, head_department, kurator, ws)
    create_kvk_style()
    wb.save('src/get_kvk/template/test.xlsx')
    wb.close()


@router.get('/get_kvk')
async def get_kvk(request: Request, user: User = Depends(current_active_user),
                  session: AsyncSession = Depends(get_async_session)):
    file_path = 'src/get_kvk/template/KVKTemplate.xlsx'
    df_risk, df_executor, head_department, kurator = await select_risks(3, session)
    make_kvk(df_risk, df_executor, head_department, kurator)

    return 202
    # return FileResponse(file_path)
